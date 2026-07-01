"""Verify a Lasagna workbook against captured full-route row expectations."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
from pathlib import Path
from typing import Any
from zipfile import ZipFile

from openpyxl import load_workbook

from lasagna.workbook.writer import FAILED_SOURCE_ROWS_TITLE

ROUTE_HEADERS = (
    "Location ID",
    "Site Code",
    "Site Type",
    "Site Type No",
    "NE Information",
    "Cabling Location",
    "Cabling Points",
    "Conn Type",
    "Location Alias",
    "PCG pos NwP Id",
    "Route Path",
    "Pos",
    "Prot",
    "Status o-time",
    "O-time",
    "Status t-time",
    "T-time",
    "Comment",
)

BASELINE_FIELDS = (
    "service_id",
    "site_code",
    "site_type",
    "site_type_no",
    "route_path",
    "pos",
    "ne_info",
    "cabling_location",
    "cabling_points",
    "connection_type",
)


def _text(value: object) -> str | None:
    text = "" if value is None else str(value)
    return text or None


def _hash_rows(rows: list[dict[str, Any]]) -> str:
    payload = json.dumps(rows, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest().upper()


def _formula_xml_count(workbook_path: Path) -> int:
    with ZipFile(workbook_path) as workbook_zip:
        return sum(
            workbook_zip.read(name).count(b"<f")
            for name in workbook_zip.namelist()
            if name.startswith("xl/worksheets/")
        )


def _combined_csv_row_count(combined_csv_path: Path) -> int:
    with combined_csv_path.open(newline="", encoding="utf-8") as handle:
        return sum(1 for _row in csv.DictReader(handle))


def _summary_rows(workbook_path: Path) -> dict[str, dict[str, Any]]:
    workbook = load_workbook(workbook_path, data_only=True)
    try:
        sheet = workbook["Summary"]
        results: dict[str, dict[str, Any]] = {}
        for values in sheet.iter_rows(min_row=2, values_only=True):
            if not values or values[2] is None:
                continue
            service_id = str(values[2])
            results[service_id] = {
                "status": values[3],
                "row_count": values[4],
                "migration_row_count": values[5],
                "route_order_source": values[8],
                "message": values[9],
            }
        return results
    finally:
        workbook.close()


def _service_route_rows(workbook_path: Path, service_id: str) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, data_only=True)
    try:
        if service_id not in workbook.sheetnames:
            return []
        sheet = workbook[service_id]
        rows: list[dict[str, Any]] = []
        row_number = 1
        while row_number <= sheet.max_row:
            values = [sheet.cell(row=row_number, column=column).value for column in range(1, 19)]
            if tuple(values) != ROUTE_HEADERS:
                row_number += 1
                continue
            title = sheet.cell(row=row_number - 1, column=1).value if row_number > 1 else None
            target_rows = [] if title == FAILED_SOURCE_ROWS_TITLE else rows
            row_number = _append_route_section_rows(sheet, row_number + 1, target_rows, service_id)
        return rows
    finally:
        workbook.close()


def _append_route_section_rows(
    sheet: Any, row_number: int, rows: list[dict[str, Any]], service_id: str
) -> int:
    while row_number <= sheet.max_row:
        values = [sheet.cell(row=row_number, column=column).value for column in range(1, 19)]
        if tuple(values) == ROUTE_HEADERS:
            return row_number
        route_values: dict[str, object] = {
            header: sheet.cell(row=row_number, column=column).value
            for column, header in enumerate(ROUTE_HEADERS, start=1)
        }
        if _is_route_section_terminator(route_values):
            return row_number + 1
        rows.append(
            {
                "service_id": service_id,
                "site_code": _text(route_values["Site Code"]),
                "site_type": _text(route_values["Site Type"]),
                "site_type_no": _text(route_values["Site Type No"]),
                "route_path": _text(route_values["Route Path"]),
                "pos": _text(route_values["Pos"]),
                "ne_info": _text(route_values["NE Information"]),
                "cabling_location": _text(route_values["Cabling Location"]),
                "cabling_points": _text(route_values["Cabling Points"]),
                "connection_type": _text(route_values["Conn Type"]),
            }
        )
        row_number += 1
    return row_number


def _is_route_section_terminator(route_values: dict[str, object]) -> bool:
    marker_values = {"No route rows found.", "No migration portion found."}
    if any(value in marker_values for value in route_values.values()):
        return True
    return all(value in (None, "") for value in route_values.values())


def _expected_rows(expected_service: dict[str, Any], service_id: str) -> list[dict[str, Any]]:
    rows = [
        row
        for row in expected_service.get("rows", [])
        if row.get("site_code") and row.get("route_path") and row.get("site_code") != "Site Code"
    ]
    return [
        {
            field: row.get(field) if field != "service_id" else service_id
            for field in BASELINE_FIELDS
        }
        for row in rows
    ]


def capture_expectations(workbook_path: Path) -> dict[str, Any]:
    """Capture current workbook route rows as a full-route expectation file."""
    summary = _summary_rows(workbook_path)
    services: dict[str, dict[str, Any]] = {}
    for service_id, summary_row in summary.items():
        route_rows = _service_route_rows(workbook_path, service_id)
        services[service_id] = {
            "status": summary_row.get("status"),
            "route_order_source": summary_row.get("route_order_source"),
            "message": summary_row.get("message"),
            "rows": [
                {field: row.get(field) for field in BASELINE_FIELDS if field != "service_id"}
                for row in route_rows
            ],
        }
    return {
        "source_workbook": str(workbook_path),
        "formula_xml_count": _formula_xml_count(workbook_path),
        "services": services,
    }


FAIL_CLOSED_MESSAGE_FRAGMENTS = (
    "DP/SDP endpoint role not proven",
    "missing node/media fact(s)",
    "missing route contract",
    "missing SITE_SIDE",
    "conflicts with",
    "not an endpoint",
    "duplicate",
    "collapsed edge lacks explicit side proof",
    "migration route contract not proven",
    "transport adjacency path not uniquely proven",
    "transport adjacency path not proven",
    "device transport endpoint not proven",
)


def _is_valid_fail_closed(
    expected_service: dict[str, Any],
    summary_row: dict[str, Any],
    actual_rows: list[dict[str, Any]],
) -> bool:
    if expected_service.get("status") != "SORT FAILED":
        return False
    if summary_row.get("status") != "SORT FAILED":
        return False
    if actual_rows:
        return False
    message = str(summary_row.get("message") or "")
    return any(fragment in message for fragment in FAIL_CLOSED_MESSAGE_FRAGMENTS)


def _first_difference(
    expected_rows: list[dict[str, Any]],
    actual_rows: list[dict[str, Any]],
) -> dict[str, Any] | None:
    for index, expected_row in enumerate(expected_rows):
        if index >= len(actual_rows):
            return {"row_index": index, "expected": expected_row, "actual": None}
        actual_row = actual_rows[index]
        if actual_row != expected_row:
            return {"row_index": index, "expected": expected_row, "actual": actual_row}
    if len(actual_rows) > len(expected_rows):
        return {
            "row_index": len(expected_rows),
            "expected": None,
            "actual": actual_rows[len(expected_rows)],
        }
    return None


def verify_workbook(
    workbook_path: Path,
    combined_csv_path: Path,
    expected_json_path: Path,
) -> dict[str, Any]:
    expected = json.loads(expected_json_path.read_text(encoding="utf-8"))
    summary = _summary_rows(workbook_path)
    verifier_failures: list[dict[str, Any]] = []
    services: list[dict[str, Any]] = []

    for service_id, expected_service in expected["services"].items():
        expected_service_rows = _expected_rows(expected_service, service_id)
        actual_service_rows = _service_route_rows(workbook_path, service_id)
        summary_row = summary.get(service_id, {})
        first_difference = _first_difference(expected_service_rows, actual_service_rows)
        valid_ok = (
            not first_difference
            and expected_service.get("status") == "OK"
            and summary_row.get("status") == "OK"
            and summary_row.get("route_order_source") == "STRUCTURED_ROUTE_CONTRACT"
        )
        valid_fail_closed = _is_valid_fail_closed(
            expected_service, summary_row, actual_service_rows
        )
        service_status = "OK" if valid_ok or valid_fail_closed else "FAIL"
        service_report = {
            "service_id": service_id,
            "status": service_status,
            "workbook_status": summary_row.get("status"),
            "route_order_source": summary_row.get("route_order_source"),
            "message": summary_row.get("message"),
            "expected_row_count": len(expected_service_rows),
            "actual_row_count": len(actual_service_rows),
            "expected_route_hash": _hash_rows(expected_service_rows),
            "actual_route_hash": _hash_rows(actual_service_rows),
            "first_difference": first_difference,
        }
        services.append(service_report)
        if service_status != "OK":
            verifier_failures.append(service_report)

    formula_xml_count = _formula_xml_count(workbook_path)
    raw_csv_cleanup_count = 0
    if formula_xml_count:
        verifier_failures.append(
            {
                "service_id": None,
                "status": "FAIL",
                "message": f"formula_xml_count={formula_xml_count}",
            }
        )

    return {
        "status": "OK" if not verifier_failures else "FAIL",
        "workbook": str(workbook_path),
        "combined_csv": str(combined_csv_path),
        "expected_json": str(expected_json_path),
        "formula_xml_count": formula_xml_count,
        "raw_csv_cleanup_count": raw_csv_cleanup_count,
        "combined_csv_row_count": _combined_csv_row_count(combined_csv_path),
        "services": services,
        "verifier_failures": verifier_failures,
    }


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--capture-expected",
        action="store_true",
        help="write expected full-route rows from --workbook to --output",
    )
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--combined-csv", type=Path)
    parser.add_argument("--expected-json", type=Path)
    parser.add_argument("--output", type=Path, required=True)
    return parser.parse_args()


def main() -> int:
    args = _parse_args()
    if args.capture_expected:
        result = capture_expectations(args.workbook)
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(json.dumps(result, indent=2, sort_keys=True), encoding="utf-8")
        print("status=OK")
        print(f"captured_services={len(result['services'])}")
        print(f"formula_xml_count={result['formula_xml_count']}")
        print(f"output={args.output}")
        return 0
    if args.combined_csv is None or args.expected_json is None:
        raise SystemExit(
            "--combined-csv and --expected-json are required unless --capture-expected is used"
        )
    result = verify_workbook(args.workbook, args.combined_csv, args.expected_json)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(result, indent=2, sort_keys=True), encoding="utf-8")
    print(f"status={result['status']}")
    print(f"formula_xml_count={result['formula_xml_count']}")
    print(f"raw_csv_cleanup_count={result['raw_csv_cleanup_count']}")
    print(f"verifier_failures={len(result['verifier_failures'])}")
    print(f"output={args.output}")
    return 0 if result["status"] == "OK" else 1


if __name__ == "__main__":
    raise SystemExit(main())
