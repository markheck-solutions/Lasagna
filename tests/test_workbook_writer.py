from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook

from lasagna.domain.route_models import ROUTE_COLUMNS, RouteRow, ServiceRouteResult
from lasagna.domain.service_ids import parse_service_id_text
from lasagna.workbook.writer import (
    FAILED_SOURCE_ROWS_TITLE,
    NO_MIGRATION_MESSAGE,
    NO_ROUTE_ROWS_MESSAGE,
    SUMMARY_COLUMNS,
    write_route_workbooks,
)


def _route_row(site_code: str, route_path: str, pos: str = "1") -> RouteRow:
    return RouteRow(
        location_id=f"LOC-{site_code}",
        site_code=site_code,
        site_type="XS",
        site_type_no="107",
        ne_information=f"{site_code} router",
        cabling_location=f"[BLDG]{site_code}/01/RU01/.",
        cabling_points=f"{pos} Cable.{pos}",
        conn_type="LC",
        location_alias=f"{site_code} alias",
        pcg_pos_nwp_id=f"NWP-{site_code}",
        route_path=route_path,
        pos=pos,
        prot="N",
        status_o_time="",
        o_time="",
        status_t_time="",
        t_time="",
        comment="synthetic",
    )


def _summary_rows(path: Path) -> list[list[object]]:
    workbook = load_workbook(path, data_only=True)
    try:
        sheet = workbook["Summary"]
        return [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def test_writer_splits_workbooks_and_records_every_input(tmp_path: Path) -> None:
    text = " ".join(f"IC-{index:06d}" for index in range(1, 106))
    parsed = parse_service_id_text(f"{text} invalid IC-000001")
    service_results = {
        f"IC-{index:06d}": ServiceRouteResult.ok(
            f"IC-{index:06d}",
            (_route_row("AAA", "AAA-BBB OL01"),),
        )
        for index in range(1, 106)
    }

    results = write_route_workbooks(parsed, service_results, tmp_path, max_service_tabs=100)

    assert [result.path.name for result in results] == [
        "Lasagna_Batch_001.xlsx",
        "Lasagna_Batch_002.xlsx",
    ]
    assert len(results[0].service_ids) == 100
    assert len(results[1].service_ids) == 5

    first_summary = _summary_rows(results[0].path)
    second_summary = _summary_rows(results[1].path)
    assert first_summary[0] == list(SUMMARY_COLUMNS)
    assert len(first_summary) == 103
    assert len(second_summary) == 6
    assert first_summary[-2][3] == "INVALID ID"
    assert first_summary[-1][9] == "Duplicate of input 1."


def test_writer_uses_exact_route_columns_and_migration_section(tmp_path: Path) -> None:
    parsed = parse_service_id_text("ICB-222222")
    service_results = {
        "ICB-222222": ServiceRouteResult.ok(
            "ICB-222222",
            (_route_row("ALPHA", "ALPHA-BETA OL01"),),
            (_route_row("BETA", "ALPHA-BETA OL01", "2"),),
            route_order_source="ROUTE_ORDER_METADATA",
        )
    }

    [result] = write_route_workbooks(parsed, service_results, tmp_path)
    workbook = load_workbook(result.path, data_only=True)
    try:
        sheet = workbook["ICB-222222"]
        assert sheet["A5"].value == "Sorted Route Path"
        assert [cell.value for cell in sheet[6]] == list(ROUTE_COLUMNS)
        assert sheet["A9"].value == "Migration Portion"
        assert [cell.value for cell in sheet[10]] == list(ROUTE_COLUMNS)
        assert sheet["A11"].value == "LOC-BETA"
        assert workbook["Summary"]["I2"].value == "ROUTE_ORDER_METADATA"
    finally:
        workbook.close()


def test_writer_keeps_no_migration_note_on_same_service_tab(tmp_path: Path) -> None:
    parsed = parse_service_id_text("IC-333333")
    service_results = {
        "IC-333333": ServiceRouteResult.ok(
            "IC-333333",
            (_route_row("ALPHA", "ALPHA-BETA OL01"),),
        )
    }

    [result] = write_route_workbooks(parsed, service_results, tmp_path)
    workbook = load_workbook(result.path, data_only=True)
    try:
        sheet = workbook["IC-333333"]
        assert sheet["A9"].value == "Migration Portion"
        assert sheet["A11"].value == NO_MIGRATION_MESSAGE
    finally:
        workbook.close()


def test_writer_keeps_service_failures_isolated(tmp_path: Path) -> None:
    parsed = parse_service_id_text("IC-444444 ICB-555555 IC-666666")
    service_results = {
        "IC-444444": ServiceRouteResult.ok(
            "IC-444444",
            (_route_row("ALPHA", "ALPHA-BETA OL01"),),
        ),
        "ICB-555555": ServiceRouteResult.sort_failed(
            "ICB-555555",
            "ROUTE_ORDER_METADATA completeness partial for ICB-555555.",
        ),
    }

    [result] = write_route_workbooks(parsed, service_results, tmp_path)
    workbook = load_workbook(result.path, data_only=True)
    try:
        assert set(workbook.sheetnames) == {"Summary", "IC-444444", "ICB-555555", "IC-666666"}
        assert workbook["ICB-555555"]["B2"].value == "SORT FAILED"
        assert workbook["ICB-555555"]["A7"].value == NO_ROUTE_ROWS_MESSAGE
        assert workbook["IC-666666"]["B2"].value == "NO DATA"
        assert workbook["IC-666666"]["A7"].value == NO_ROUTE_ROWS_MESSAGE
        summary_statuses = [workbook["Summary"][f"D{row}"].value for row in range(2, 5)]
        assert summary_statuses == ["OK", "SORT FAILED", "NO DATA"]
    finally:
        workbook.close()


def test_writer_labels_failed_visible_rows_as_source_rows_not_route_proof(
    tmp_path: Path,
) -> None:
    parsed = parse_service_id_text("IC-888888")
    service_results = {
        "IC-888888": ServiceRouteResult.sort_failed(
            "IC-888888",
            "transport adjacency path not proven. Source rows shown for troubleshooting only; not route proof.",
            (_route_row("ALPHA", "ALPHA-BETA OL01"),),
            route_order_source="SOURCE_ROWS_NOT_ROUTE_PROOF",
        )
    }

    [result] = write_route_workbooks(parsed, service_results, tmp_path)
    workbook = load_workbook(result.path, data_only=True)
    try:
        service_sheet = workbook["IC-888888"]
        summary = workbook["Summary"]
        assert service_sheet["B2"].value == "SORT FAILED"
        assert service_sheet["A5"].value == FAILED_SOURCE_ROWS_TITLE
        assert service_sheet["A7"].value == "LOC-ALPHA"
        assert summary["D2"].value == "SORT FAILED"
        assert summary["E2"].value == 1
        assert summary["I2"].value == "SOURCE_ROWS_NOT_ROUTE_PROOF"
    finally:
        workbook.close()


def test_writer_keeps_route_text_literals_out_of_formula_xml(tmp_path: Path) -> None:
    parsed = parse_service_id_text("IC-777777")
    special_rows = tuple(
        _route_row("ALPHA", route_path, str(index))
        for index, route_path in enumerate(("=A1", "+PLUS", "-MINUS", "@AT"), start=1)
    )
    service_results = {
        "IC-777777": ServiceRouteResult.ok(
            "IC-777777",
            special_rows,
            route_order_source="=SOURCE",
            message="=BEARER",
        )
    }

    [result] = write_route_workbooks(parsed, service_results, tmp_path)

    workbook = load_workbook(result.path, data_only=False)
    try:
        sheet = workbook["IC-777777"]
        assert [sheet[f"K{row}"].value for row in range(7, 11)] == [
            "=A1",
            "+PLUS",
            "-MINUS",
            "@AT",
        ]
        assert workbook["Summary"]["I2"].value == "=SOURCE"
        assert workbook["Summary"]["J2"].value == "=BEARER"
    finally:
        workbook.close()

    with ZipFile(result.path) as workbook_zip:
        worksheet_xml = [
            workbook_zip.read(name)
            for name in workbook_zip.namelist()
            if name.startswith("xl/worksheets/")
        ]
    assert all(b"<f" not in xml for xml in worksheet_xml)
