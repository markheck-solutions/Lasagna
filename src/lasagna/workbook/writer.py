"""Write Lasagna route review workbooks."""

from __future__ import annotations

from collections.abc import Iterable
from dataclasses import dataclass
from pathlib import Path
from typing import cast

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from lasagna.domain.route_models import ROUTE_COLUMNS, RouteRow, ServiceRouteResult
from lasagna.domain.service_ids import ParsedServiceInput, unique_valid_service_ids
from lasagna.workbook.names import sanitize_sheet_name, workbook_filename

SUMMARY_COLUMNS: tuple[str, ...] = (
    "Input Order",
    "Input Text",
    "Normalized ID",
    "Status",
    "Rows",
    "Migration Rows",
    "Workbook",
    "Sheet",
    "Route Order Source",
    "Message",
)

INVALID_ID_MESSAGE = "Invalid service ID; expected IC-123456 or ICB-123456."
NO_MIGRATION_MESSAGE = "No migration portion found."
NO_ROUTE_ROWS_MESSAGE = "No route rows found."
FAILED_SOURCE_ROWS_TITLE = "Source Route Rows (Not Route Proof)"
CellValue = str | int | None


@dataclass(frozen=True)
class WorkbookWriteResult:
    """One generated workbook and the unique service IDs assigned to it."""

    path: Path
    service_ids: tuple[str, ...]


@dataclass(frozen=True)
class WorkbookServicePlacement:
    """Workbook/sheet assignment for one unique valid service ID."""

    workbook_index: int
    workbook_name: str
    sheet_name: str


def _chunks(items: list[str], size: int) -> list[tuple[str, ...]]:
    if size < 1:
        raise ValueError("max_service_tabs must be 1 or greater")
    return [tuple(items[index : index + size]) for index in range(0, len(items), size)]


def _planned_workbooks(service_ids: list[str], max_service_tabs: int) -> list[tuple[str, ...]]:
    chunks = _chunks(service_ids, max_service_tabs)
    return chunks or [()]


def _service_result(
    service_id: str,
    service_results: dict[str, ServiceRouteResult],
) -> ServiceRouteResult:
    return service_results.get(service_id, ServiceRouteResult.no_data(service_id))


def _write_cell_value(ws: Worksheet, row_number: int, column_number: int, value: CellValue) -> None:
    cell = ws.cell(row=row_number, column=column_number)
    cell.value = value
    if isinstance(value, str):
        cell.data_type = "s"


def _write_row_values(ws: Worksheet, row_number: int, values: Iterable[CellValue]) -> None:
    for column_number, value in enumerate(values, start=1):
        _write_cell_value(ws, row_number, column_number, value)


def _write_route_rows(ws: Worksheet, row_number: int, rows: tuple[RouteRow, ...]) -> int:
    for route_row in rows:
        _write_row_values(ws, row_number, route_row.values())
        row_number += 1
    return row_number


def _write_route_section(
    ws: Worksheet,
    row_number: int,
    title: str,
    rows: tuple[RouteRow, ...],
    empty_message: str,
) -> int:
    _write_cell_value(ws, row_number, 1, title)
    row_number += 1
    _write_row_values(ws, row_number, ROUTE_COLUMNS)
    row_number += 1
    if rows:
        return _write_route_rows(ws, row_number, rows)
    _write_cell_value(ws, row_number, 1, empty_message)
    return row_number + 1


def _write_service_sheet(
    ws: Worksheet,
    service_result: ServiceRouteResult,
) -> None:
    _write_cell_value(ws, 1, 1, "Service ID")
    _write_cell_value(ws, 1, 2, service_result.service_id)
    _write_cell_value(ws, 2, 1, "Status")
    _write_cell_value(ws, 2, 2, service_result.status)
    _write_cell_value(ws, 3, 1, "Message")
    _write_cell_value(ws, 3, 2, service_result.message)

    route_title = (
        FAILED_SOURCE_ROWS_TITLE
        if service_result.status == "SORT FAILED" and service_result.sorted_rows
        else "Sorted Route Path"
    )
    row_number = _write_route_section(
        ws,
        5,
        route_title,
        service_result.sorted_rows,
        NO_ROUTE_ROWS_MESSAGE,
    )
    row_number += 1
    _write_route_section(
        ws,
        row_number,
        "Migration Portion",
        service_result.migration_rows,
        NO_MIGRATION_MESSAGE,
    )


def _build_placements(
    workbook_services: list[tuple[str, ...]],
) -> dict[str, WorkbookServicePlacement]:
    placements: dict[str, WorkbookServicePlacement] = {}
    for workbook_index, service_ids in enumerate(workbook_services):
        workbook_name = workbook_filename(workbook_index + 1)
        existing_sheet_names = {"Summary"}
        for service_id in service_ids:
            sheet_name = sanitize_sheet_name(service_id, existing_sheet_names)
            existing_sheet_names.add(sheet_name)
            placements[service_id] = WorkbookServicePlacement(
                workbook_index=workbook_index,
                workbook_name=workbook_name,
                sheet_name=sheet_name,
            )
    return placements


def _summary_inputs_for_workbook(
    parsed_inputs: list[ParsedServiceInput],
    placements: dict[str, WorkbookServicePlacement],
    workbook_index: int,
) -> list[ParsedServiceInput]:
    summary_inputs: list[ParsedServiceInput] = []
    for parsed_input in parsed_inputs:
        if not parsed_input.is_valid:
            if workbook_index == 0:
                summary_inputs.append(parsed_input)
            continue
        placement = placements.get(parsed_input.normalized_id)
        if placement and placement.workbook_index == workbook_index:
            summary_inputs.append(parsed_input)
    return summary_inputs


def _summary_row(
    parsed_input: ParsedServiceInput,
    placements: dict[str, WorkbookServicePlacement],
    service_results: dict[str, ServiceRouteResult],
) -> tuple[CellValue, ...]:
    if not parsed_input.is_valid:
        first_workbook = workbook_filename(1)
        return (
            parsed_input.input_order,
            parsed_input.input_text,
            parsed_input.normalized_id,
            "INVALID ID",
            0,
            0,
            first_workbook,
            "",
            "",
            INVALID_ID_MESSAGE,
        )

    placement = placements[parsed_input.normalized_id]
    result = _service_result(parsed_input.normalized_id, service_results)
    message = result.message
    if parsed_input.duplicate_of is not None:
        message = f"Duplicate of input {parsed_input.duplicate_of}."
    return (
        parsed_input.input_order,
        parsed_input.input_text,
        parsed_input.normalized_id,
        result.status,
        len(result.sorted_rows),
        len(result.migration_rows),
        placement.workbook_name,
        placement.sheet_name,
        result.route_order_source,
        message,
    )


def _write_summary_sheet(
    ws: Worksheet,
    summary_inputs: list[ParsedServiceInput],
    placements: dict[str, WorkbookServicePlacement],
    service_results: dict[str, ServiceRouteResult],
) -> None:
    ws.title = "Summary"
    _write_row_values(ws, 1, SUMMARY_COLUMNS)
    for row_number, parsed_input in enumerate(summary_inputs, start=2):
        _write_row_values(ws, row_number, _summary_row(parsed_input, placements, service_results))


def write_route_workbooks(
    parsed_inputs: list[ParsedServiceInput],
    service_results: dict[str, ServiceRouteResult],
    output_dir: Path,
    max_service_tabs: int = 100,
) -> list[WorkbookWriteResult]:
    """Write split Lasagna route workbooks for parsed inputs."""
    output_dir.mkdir(parents=True, exist_ok=True)
    service_ids = unique_valid_service_ids(parsed_inputs)
    workbook_services = _planned_workbooks(service_ids, max_service_tabs)
    placements = _build_placements(workbook_services)
    results: list[WorkbookWriteResult] = []

    for workbook_index, workbook_service_ids in enumerate(workbook_services):
        workbook_path = output_dir / workbook_filename(workbook_index + 1)
        workbook = Workbook()
        try:
            summary_inputs = _summary_inputs_for_workbook(
                parsed_inputs,
                placements,
                workbook_index,
            )
            _write_summary_sheet(
                cast(Worksheet, workbook.active),
                summary_inputs,
                placements,
                service_results,
            )
            for service_id in workbook_service_ids:
                placement = placements[service_id]
                worksheet = workbook.create_sheet(title=placement.sheet_name)
                _write_service_sheet(worksheet, _service_result(service_id, service_results))
            workbook.save(workbook_path)
        finally:
            workbook.close()
        results.append(WorkbookWriteResult(path=workbook_path, service_ids=workbook_service_ids))

    return results
