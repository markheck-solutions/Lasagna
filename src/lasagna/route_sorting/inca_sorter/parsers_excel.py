"""Excel export parsing for INCA route rows."""

from __future__ import annotations

import re
import sys
from collections import defaultdict
from typing import Any, cast

import openpyxl

from .models import INCA_COLUMNS, InCARow
from .parsers_common import _safe_int, _safe_str


def extract_service_id(filepath: str) -> str | None:
    """Extract service ID from Excel file title row.

    INCA exports have a title in row 1, cell A1:
    'Route path with cabling points - IC-136025 | ...'
    'Route path with cabling points - ICB-811386 | ...'

    Returns the service ID (e.g., 'IC-136025', 'ICB-811386') or None.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = cast(Any, wb.active)
    row1 = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    wb.close()
    if not row1 or not row1[0]:
        return None
    title = str(row1[0][0]) if row1[0][0] else ""
    m = re.search(r"(ICB?-\d+)", title)
    return m.group(1) if m else None


def _map_excel_headers_case_insensitively(
    row_cells: tuple[Any, ...],
    inca_col_lower: dict[str, str],
) -> dict[str, int]:
    """Map case-insensitive Excel headers to canonical INCA column names."""
    header_map: dict[str, int] = {}
    for cell in row_cells:
        header = _safe_str(cell.value).lower()
        if header in inca_col_lower:
            header_map[inca_col_lower[header]] = cell.column - 1
    return header_map


def _row_contains_site_code_header(row_cells: tuple[Any, ...]) -> bool:
    """Return True when a worksheet row includes the Site Code header."""
    return any(_safe_str(cell.value).lower() == "site code" for cell in row_cells)


def _fallback_excel_header_map(ws: Any) -> tuple[int, dict[str, int]]:
    """Fallback to row 1 headers for synthetic workbook fixtures."""
    header_map: dict[str, int] = {}
    first_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=False))
    if first_row and first_row[0]:
        for cell in first_row[0]:
            value = _safe_str(cell.value)
            if value in INCA_COLUMNS:
                header_map[value] = cell.column - 1
    return 1, header_map


def _find_excel_header_row(ws: Any, inca_col_lower: dict[str, str]) -> tuple[int, dict[str, int]]:
    """Find the Excel header row and its canonical INCA column map."""
    for row_num in range(1, 11):
        row_cells = list(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=False))
        if not row_cells or not row_cells[0]:
            continue
        cells = row_cells[0]
        if _row_contains_site_code_header(cells):
            return row_num, _map_excel_headers_case_insensitively(cells, inca_col_lower)
    return _fallback_excel_header_map(ws)


def _warn_missing_excel_columns(header_map: dict[str, int]) -> None:
    """Emit the existing owner-readable warning for absent INCA columns."""
    missing = set(INCA_COLUMNS) - set(header_map)
    if missing:
        print(f"WARNING: Missing columns in Excel: {missing}", file=sys.stderr)


def _excel_cell_value(row_data: tuple, header_map: dict[str, int], col_name: str) -> str:
    """Return a normalized Excel cell value for a canonical INCA column."""
    column_index = header_map.get(col_name)
    if column_index is None or column_index >= len(row_data):
        return ""
    return _safe_str(row_data[column_index])


def _make_excel_row(row_data: tuple, row_index: int, header_map: dict[str, int]) -> InCARow:
    """Create an InCARow from one Excel worksheet data row."""
    pos_value = row_data[header_map["Pos"]] if "Pos" in header_map else None
    return InCARow(
        site_code=_excel_cell_value(row_data, header_map, "Site Code"),
        site_type=_excel_cell_value(row_data, header_map, "Site Type"),
        ne_info=_excel_cell_value(row_data, header_map, "NE Information") or None,
        cabling_location=_excel_cell_value(row_data, header_map, "Cabling Location"),
        cabling_points=_excel_cell_value(row_data, header_map, "Cabling Points"),
        conn_type=_excel_cell_value(row_data, header_map, "Conn type"),
        location_alias=_excel_cell_value(row_data, header_map, "Location Alias") or None,
        route_path=_excel_cell_value(row_data, header_map, "Route Path"),
        pos=_safe_int(pos_value),
        status_o_time=_excel_cell_value(row_data, header_map, "Status o-time") or None,
        o_time=_excel_cell_value(row_data, header_map, "O-time") or None,
        status_t_time=_excel_cell_value(row_data, header_map, "Status t-time") or None,
        t_time=_excel_cell_value(row_data, header_map, "T-time") or None,
        comment=_excel_cell_value(row_data, header_map, "Comment") or None,
        row_index=row_index,
    )


def _read_excel_rows(ws: Any, data_start: int, header_map: dict[str, int]) -> list[InCARow]:
    """Read all non-blank Excel data rows into InCARow objects."""
    rows: list[InCARow] = []
    for row_index, row_data in enumerate(
        ws.iter_rows(min_row=data_start, values_only=True), start=data_start
    ):
        if all(value is None for value in row_data):
            continue
        row = _make_excel_row(row_data, row_index, header_map)
        if row.site_code:
            rows.append(row)
    return rows


def read_excel(filepath: str) -> list[InCARow]:
    """Read INCA export from Excel file, return list of InCARow objects.

    Handles real INCA exports where headers may not be on row 1.
    Scans rows 1-10 for the header row (case-insensitive column matching).
    Extra columns in the export are ignored gracefully.

    Args:
        filepath: Path to .xlsx file with INCA route path export.

    Returns:
        List of InCARow objects, one per data row.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    try:
        ws = cast(Any, wb.active)
        inca_col_lower = {col.lower(): col for col in INCA_COLUMNS}
        header_row_num, header_map = _find_excel_header_row(ws, inca_col_lower)
        _warn_missing_excel_columns(header_map)
        return _read_excel_rows(ws, header_row_num + 1, header_map)
    finally:
        wb.close()


def build_tl_device_map(
    tl_device_records: list[dict],
    service_id: str,
) -> dict[tuple[str, str], dict[str, list[str]]]:
    """Build a mapping from (service_id, site_code) to {tl_name: [ne_parts]}.

    This maps transport link names to the specific device(s) (NE_PART) that
    terminate them at each site, enabling data-driven within-site ordering.
    A single TL can have multiple NE_PARTs at one site (e.g., both XT-05 and
    OTC-02 at the same site for the same transport link).

    Args:
        tl_device_records: Raw dicts from TL_DEVICE rows in combined CSV.
        service_id: Filter to this service only.

    Returns:
        Dict keyed by (service_id, site_code) -> {tl_name: [ne_part, ...]}.
        Example: {('IC-394531', 'ATM/2'): {'ASH/2-ATM/2 OCGX05': ['XTC-08'],
                                             'ATM/2-IPLS O600G03': ['XT-44']}}
    """
    result: dict[tuple[str, str], dict[str, list[str]]] = defaultdict(lambda: defaultdict(list))
    for rec in tl_device_records:
        sid = str(rec.get("SERVICE_ID", "")).strip()
        if sid != service_id:
            continue
        tl_name = str(rec.get("TL_NAME", "")).strip()
        site_code = str(rec.get("SITE_CODE", "")).strip()
        ne_part = str(rec.get("NE_PART", "")).strip()
        if tl_name and site_code and ne_part:
            if ne_part not in result[(sid, site_code)][tl_name]:
                result[(sid, site_code)][tl_name].append(ne_part)
    return {k: dict(v) for k, v in result.items()}
