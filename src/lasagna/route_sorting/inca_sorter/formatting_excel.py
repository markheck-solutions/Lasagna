"""Workbook writers for INCA route sorter output."""

from __future__ import annotations

from collections.abc import Sequence
from typing import Any, cast

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from .models import INCA_COLUMNS, InCARow, Ticket


def _write_cells(ws: Worksheet, row_num: int, values: Sequence[Any]) -> None:
    """Write one worksheet row of values starting at column 1."""
    for col_idx, value in enumerate(values, 1):
        ws.cell(row=row_num, column=col_idx, value=value)


def _route_sheet_row_values(item: InCARow) -> list[object]:
    """Return the legacy workbook row values for one route item."""
    points_out = item.display_points if item.display_points else item.cabling_points
    return [
        item.site_code,
        item.site_type,
        item.ne_info or "",
        item.cabling_location,
        points_out,
        item.conn_type,
        item.location_alias or "",
        item.route_path,
        item.pos,
        item.status_o_time or "",
        item.o_time or "",
        item.status_t_time or "",
        item.t_time or "",
        item.comment or "",
        item.classification,
    ]


def _write_route_data_rows(
    ws: Worksheet,
    start_row: int,
    row_items: list[InCARow],
) -> int:
    """Write route rows starting at start_row and return the next free row."""
    row_num = start_row
    for item in row_items:
        _write_cells(ws, row_num, _route_sheet_row_values(item))
        row_num += 1
    return row_num


def _write_route_headers(
    ws: Worksheet,
    row_num: int,
    headers: list[str],
) -> int:
    """Write the workbook route headers and return the next free row."""
    _write_cells(ws, row_num, headers)
    return row_num + 1


def _write_route_sheet_section(
    ws: Worksheet,
    row_num: int,
    title: str,
    headers: list[str],
    row_items: list[InCARow],
) -> int:
    """Write one labeled route section and return the next free row."""
    ws.cell(row=row_num, column=1, value=title)
    row_num = _write_route_headers(ws, row_num + 1, headers)
    return _write_route_data_rows(ws, row_num, row_items)


def _route_sheet_title(service_id: str, bearer: str) -> str:
    """Return the migration workbook title row."""
    title = f"Route path - {service_id}" if service_id else "Route path"
    return f"{title} | {bearer}" if bearer else title


def _write_sorted_route_sheet(
    ws_sorted: Worksheet,
    items: list[InCARow],
    migration_portion: list[InCARow] | None,
    service_id: str,
    bearer: str,
) -> None:
    """Write the Sorted Route Path sheet for standard and migration outputs."""
    ws_sorted.title = "Sorted Route Path"
    headers = INCA_COLUMNS + ["Classification"]

    if migration_portion is None:
        row_num = _write_route_headers(ws_sorted, 1, headers)
        _write_route_data_rows(ws_sorted, row_num, items)
        return

    ws_sorted.cell(row=1, column=1, value=_route_sheet_title(service_id, bearer))
    row_num = _write_route_sheet_section(ws_sorted, 2, "Current Route Path", headers, items)
    row_num += 1
    _write_route_sheet_section(ws_sorted, row_num, "Migration Portion", headers, migration_portion)


def _write_notations_sheet(
    ws_notes: Worksheet,
    notations: list[str],
) -> None:
    """Write the Notations sheet."""
    ws_notes.title = "Notations"
    ws_notes.cell(row=1, column=1, value="Notation")
    for row_num, note in enumerate(notations, 2):
        ws_notes.cell(row=row_num, column=1, value=note)


def _write_tickets_sheet(
    ws_tickets: Worksheet,
    tickets: list[Ticket],
) -> None:
    """Write the Tickets sheet."""
    ws_tickets.title = "Tickets"
    _write_cells(ws_tickets, 1, ["Ticket", "Variant", "Line"])

    ticket_row = 2
    for ticket in tickets:
        for ticket_line in ticket.lines:
            _write_cells(
                ws_tickets, ticket_row, [ticket.cluster_name, ticket_line.variant, ticket_line.text]
            )
            ticket_row += 1


def write_output_excel(
    filepath: str,
    items: list[InCARow],
    notations: list[str],
    tickets: list[Ticket],
    migration_portion: list[InCARow] | None = None,
    service_id: str = "",
    bearer: str = "",
) -> None:
    """Write sorted route path and tickets to an Excel file.

    For migration orders (migration_portion is not None), emits section
    headers "Current Route Path" and "Migration Portion" with column
    headers repeated for each section.

    Args:
        filepath: Output .xlsx file path.
        items: Sorted InCARow objects (Current Route Path for migration).
        notations: NOTATIONS section entries.
        tickets: Generated tickets.
        migration_portion: Migration Portion rows (None for non-migration).
        service_id: Service identifier for title row.
        bearer: Bearer route path name for title row.
    """
    wb = openpyxl.Workbook()
    try:
        _write_sorted_route_sheet(
            cast(Worksheet, wb.active),
            items,
            migration_portion,
            service_id,
            bearer,
        )
        _write_notations_sheet(wb.create_sheet(), notations)
        _write_tickets_sheet(wb.create_sheet(), tickets)
        wb.save(filepath)
    finally:
        wb.close()
