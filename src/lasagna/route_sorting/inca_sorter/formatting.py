"""Output formatting: format_sorted_route_path, format_tickets, format_notations."""

from __future__ import annotations

import re
from collections import defaultdict
from collections.abc import Callable, Sequence
from datetime import date
from typing import Any, cast

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from .models import (
    INCA_COLUMNS,
    InCARow,
    Ticket,
    TicketLine,
    _is_planned,
)
from .parsers import (
    _safe_str,
)

_NO_HUB_ASSIGNMENT_SITE_TYPES = {"X", "U"}


def _format_date(raw: str | None) -> str:
    """Format INCA date (YYYYMMDD or other) to YYYY-MM-DD."""
    if not raw:
        return "unknown"
    raw = _safe_str(raw)
    if re.match(r"^\d{8}$", raw):
        return f"{raw[:4]}-{raw[4:6]}-{raw[6:8]}"
    return raw


def _site_requires_hub_assignment(site_types: set[str]) -> bool:
    """Only field-patching site types need consumable hub assignments."""
    normalized = {
        _safe_str(site_type).strip().upper()
        for site_type in site_types
        if _safe_str(site_type).strip()
    }
    if not normalized:
        return True
    return any(site_type not in _NO_HUB_ASSIGNMENT_SITE_TYPES for site_type in normalized)


def _days_until(raw: str | None) -> str:
    """Calculate days from today until the given date."""
    if not raw:
        return "?"
    raw_str = _safe_str(raw)
    try:
        if re.match(r"^\d{8}$", raw_str):
            target = date(int(raw_str[:4]), int(raw_str[4:6]), int(raw_str[6:8]))
        elif re.match(r"^\d{4}-\d{2}-\d{2}$", raw_str):
            target = date(int(raw_str[:4]), int(raw_str[5:7]), int(raw_str[8:10]))
        else:
            return "?"
        delta = (target - date.today()).days
        return str(delta)
    except (ValueError, TypeError):
        return "?"


def _append_planned_notations(
    notations: list[str],
    rows: list[InCARow],
    *,
    include_planned: bool,
) -> None:
    """Append PLANNED notation lines using legacy device and dedupe rules."""
    if not include_planned:
        return

    seen_planned: set[str] = set()
    for row in rows:
        if not _is_planned(row.status_o_time):
            continue
        device = re.sub(r":(?:Tx|Rx)\)$", "", row.ne_info or "") or "ODF"
        key = f"{device}@{row.site_code}"
        if key in seen_planned:
            continue
        seen_planned.add(key)
        date_str = _format_date(row.o_time)
        days = _days_until(row.o_time)
        notations.append(
            f"PLANNED: {device} at {row.site_code} scheduled for {date_str} "
            f"({days} days). Not yet physically patched."
        )


def _build_bo_fiber_lookup(bo_fibers: list[dict] | None) -> dict[str, list[dict]]:
    """Return BO_FIBERS keyed by lowercased A-end NE name."""
    fiber_lookup: dict[str, list[dict]] = defaultdict(list)
    for fiber in bo_fibers or []:
        ne_name = str(fiber.get("A_NE_NAME", "")).strip()
        if not ne_name:
            continue
        ne_key = ne_name.split(",")[0].strip().lower()
        fiber_lookup[ne_key].append(fiber)
    return dict(fiber_lookup)


def _extract_bug_device_name(ne_info: str | None) -> str:
    """Extract the legacy device label used in INCA bug notations."""
    if not ne_info:
        return ""
    return ne_info.split(" -")[0].strip() if " -" in ne_info else ne_info


def _extract_bo_fiber_port_target(ne_info: str | None) -> tuple[str | None, str | None]:
    """Extract the slot and sub-port used for BO_FIBERS correlation."""
    port_match = re.search(r"-\(([^)]+)\)", ne_info or "")
    if not port_match:
        return None, None

    port_str = re.sub(r"\.\d*:(?:Tx|Rx)$", "", port_match.group(1))
    if "\\" not in port_str:
        return port_str, None

    slot, sub_port = port_str.split("\\", 1)
    return slot, sub_port.split(".")[0] if sub_port else None


def _matched_bo_fiber_rows(
    fiber_records: list[dict],
    port_slot: str | None,
    port_sub: str | None,
) -> list[dict]:
    """Return exact BO_FIBERS matches for the legacy slot or sub-port rules."""
    if not port_slot or not port_sub:
        return []

    target = f"{port_slot}:{port_sub}"
    matched: list[dict] = []
    for fiber in fiber_records:
        a_location = str(fiber.get("A_LOCATION", "")).strip()
        colon_pos = a_location.find("::")
        if colon_pos < 0:
            continue
        port_section = a_location[colon_pos + 2 :]
        if port_section == target or port_section.startswith(target + ":"):
            matched.append(fiber)
    return matched


def _build_converged_bo_fiber_hint(
    fiber_records: list[dict],
    port_slot: str | None,
    port_sub: str | None,
) -> str | None:
    """Return the optional BO_FIBERS diagnostic hint when matches converge."""
    matched = _matched_bo_fiber_rows(fiber_records, port_slot, port_sub)
    if not matched or not port_slot or not port_sub:
        return None

    b_locations = {
        str(match.get("B_LOCATION", "")).strip() for match in matched if match.get("B_LOCATION")
    }
    if len(b_locations) != 1:
        return None

    b_location = next(iter(b_locations))
    b_ports = sorted(
        {
            str(match.get("B_CONNECTION_POINT", "")).strip().split()[0]
            for match in matched
            if str(match.get("B_CONNECTION_POINT", "")).strip()
        }
    )
    full_port = f"{port_slot}/{port_sub}"
    hint = f"*Previous {full_port} was on {b_location}"
    if len(b_ports) == 1:
        return f"{hint} port {b_ports[0]}"
    if len(b_ports) == 2:
        return f"{hint} ports {b_ports[0]}+{b_ports[1]}"
    if len(b_ports) > 2:
        return f"{hint} ports {b_ports[0]}+{b_ports[-1]}"
    return hint


def _append_inca_bug_notations(
    notations: list[str],
    rows: list[InCARow],
    bo_fibers: list[dict] | None,
) -> None:
    """Append VERIFY notations for NE-location and optional BO_FIBERS hints."""
    bo_fiber_lookup = _build_bo_fiber_lookup(bo_fibers)
    seen_bugs: set[str] = set()
    for row in rows:
        if not row.has_inca_bug:
            continue

        device = _extract_bug_device_name(row.ne_info)
        key = f"{device}@{row.site_code}"
        if key in seen_bugs:
            continue
        seen_bugs.add(key)

        notation = f"VERIFY: {device} has NE-Location, not BO ODF."
        ne_name = (row.ne_info or "").strip().split(" ")[0].lower()
        port_slot, port_sub = _extract_bo_fiber_port_target(row.ne_info)
        hint = _build_converged_bo_fiber_hint(
            bo_fiber_lookup.get(ne_name, []),
            port_slot,
            port_sub,
        )
        if hint:
            notation += f"\n    {hint}"
        notations.append(notation)


def _append_decommission_notations(
    notations: list[str],
    rows: list[InCARow],
    *,
    include_decommission: bool,
) -> None:
    """Append DECOMMISSION notation lines using legacy dedupe rules."""
    if not include_decommission:
        return

    seen_decom: set[str] = set()
    for row in rows:
        if not _is_planned(row.status_t_time):
            continue
        device = re.sub(r":(?:Tx|Rx)\)$", "", row.ne_info or "") or "ODF"
        key = f"{device}@{row.site_code}"
        if key in seen_decom:
            continue
        seen_decom.add(key)
        date_str = _format_date(row.t_time)
        notations.append(
            f"DECOMMISSION: {device} at {row.site_code} scheduled for removal on {date_str}."
        )


def _build_hub_assignment_lookup(
    hub_records: list[dict] | None,
) -> dict[str, tuple[str, str]]:
    """Return normalized hub assignment metadata keyed by site code."""
    hub_lookup: dict[str, tuple[str, str]] = {}
    for record in hub_records or []:
        site_code = str(record.get("SITE_CODE", "")).strip()
        consumable = str(record.get("CONSUMABLE", "")).strip()
        hub = str(record.get("HUB", "")).strip()
        if site_code and (consumable or site_code not in hub_lookup):
            hub_lookup[site_code] = (consumable, hub)
    return hub_lookup


def _site_types_by_site(
    rows: list[InCARow],
    hub_check_sites: set[str] | None,
) -> dict[str, set[str]]:
    """Collect site types for sites that should be checked for hub assignments."""
    site_types: dict[str, set[str]] = {}
    for row in rows:
        site = _safe_str(row.site_code).strip()
        if not site:
            continue
        if hub_check_sites is not None and site not in hub_check_sites:
            continue
        site_types.setdefault(site, set()).add(row.site_type)
    return site_types


def _append_missing_hub_assignment_notations(
    notations: list[str],
    rows: list[InCARow],
    hub_records: list[dict] | None,
    hub_check_sites: set[str] | None,
) -> None:
    """Append VERIFY notations for sites missing a usable hub assignment."""
    if not hub_records:
        return

    hub_lookup = _build_hub_assignment_lookup(hub_records)
    for site in sorted(_site_types_by_site(rows, hub_check_sites)):
        site_types = _site_types_by_site(rows, hub_check_sites)[site]
        if not _site_requires_hub_assignment(site_types):
            continue
        consumable, _hub = hub_lookup.get(site, ("", ""))
        if consumable and consumable.upper() != "N/A":
            continue
        notations.append(f"VERIFY: {site} has no hub assignment. Check consumables sourcing.")


def generate_notations(
    rows: list[InCARow],
    is_pure_add: bool = False,
    trunk_edges: Sequence[tuple[str, str, str | None]] | None = None,
    is_migration: bool = False,
    hub_records: list[dict] | None = None,
    bo_fibers: list[dict] | None = None,
    hub_check_sites: set[str] | None = None,
) -> list[str]:
    """Generate NOTATIONS section entries.

    Includes:
    - PLANNED alerts for Status o-time = Planned (Mixed Add only, omitted for Pure Add
      and Migration)
    - INCA BUG alerts for NCS + NE-Location (all order types), with an optional
      exact-match BO_FIBERS diagnostic line when evidence converges on one ODF
    - DECOMMISSION alerts for Status t-time = Planned (omitted for Migration)
    - VERIFY alerts for sites without hub assignments

    For migration orders, PLANNED and DECOMMISSION are expected states and
    alerting about them is noise. VERIFY (INCA bug) remains actionable.

    Args:
        rows: All rows (sorted or unsorted).
        is_pure_add: True if all rows are Planned (Pure Add). Suppresses PLANNED alerts.
        trunk_edges: Trunk edges from parse_trunk_edges(). If provided, enables
            PP-184 suspicious trunk detection.
        is_migration: True for migration orders. Suppresses PLANNED and DECOMMISSION.
        hub_records: Optional HUB_SITE records for consumable hub notations.
        bo_fibers: Optional BO_FIBERS records for tightly gated INCA BUG
            diagnostic enrichment.
        hub_check_sites: Optional site codes that actually receive field-tech
            patching work. When provided, missing-hub VERIFY notations are
            limited to this set.

    Returns:
        List of notation strings.
    """
    notations: list[str] = []

    _append_planned_notations(
        notations,
        rows,
        include_planned=not is_pure_add and not is_migration,
    )
    _append_inca_bug_notations(notations, rows, bo_fibers)
    _append_decommission_notations(
        notations,
        rows,
        include_decommission=not is_migration,
    )
    _append_missing_hub_assignment_notations(
        notations,
        rows,
        hub_records,
        hub_check_sites,
    )

    return notations


def _format_route_rows(items: list[InCARow]) -> list[str]:
    """Format InCARow items into display lines."""
    lines: list[str] = []
    for item in items:
        marker = ""
        if item.classification == "NEW":
            marker = " [NEW]"
        elif item.classification == "DECOMMISSION":
            marker = " [DECOM]"

        ne = item.ne_info or "-"
        points_display = item.display_points if item.display_points else item.cabling_points
        lines.append(
            f"  {item.site_code:<12} {item.site_type:<4} "
            f"{ne:<25} "
            f"{item.cabling_location:<40} {points_display:<15} "
            f"Pos:{item.pos}{marker}"
        )
    return lines


def _append_ticket_lines_with_site_breaks(
    lines: list[str],
    ticket_lines: list[TicketLine],
    formatter: Callable[[TicketLine], str] | None = None,
) -> None:
    """Append formatted ticket lines with blank-line separators between cable groups.

    Each fiber cable (jumper) connects two ODF positions, producing a pair of
    consecutive ticket lines.  Blank lines separate cable pairs and site
    transitions for field crew readability.
    """
    prev_site: str | None = None
    for i, ticket_line in enumerate(ticket_lines):
        # Blank line on cable boundary (every 2 lines) or site change
        if i > 0 and (i % 2 == 0 or ticket_line.site_code != prev_site):
            lines.append("")
        if formatter is None:
            lines.append(ticket_line.text)
        else:
            lines.append(formatter(ticket_line))
        prev_site = ticket_line.site_code


def _format_stage1_ticket_line(ticket_line: TicketLine) -> str:
    """Format one Stage 1 ticket line with its inline action label."""
    label = ticket_line.hotcut_label if ticket_line.hotcut_label else ticket_line.classification
    if label in ("UNCHANGED", "LIVE"):
        return f"{ticket_line.text} *** HOT-CUT SIDE - Leave Hanging ***"
    if label in ("NEW_ONLY", "NEW"):
        return f"{ticket_line.text} *** CONNECT ***"
    return ticket_line.text


def format_sorted_route_path(
    items: list[InCARow],
    migration_portion: list[InCARow] | None = None,
) -> str:
    """Format the sorted route path for stdout display.

    For migration orders, emits "Current Route Path" and "Migration Portion"
    section headers. For standard Add orders (migration_portion is None),
    emits the single "SORTED ROUTE PATH" header.

    Args:
        items: Sorted InCARow objects (Current Route Path for migration).
        migration_portion: Migration Portion rows (None for non-migration).

    Returns:
        Formatted string with row data.
    """
    lines: list[str] = []

    if migration_portion is not None:
        # Migration order: two sections with headers
        lines.append("=" * 80)
        lines.append("Current Route Path")
        lines.append("=" * 80)
        lines.append("")
        lines.extend(_format_route_rows(items))
        lines.append("")
        lines.append("")
        lines.append("=" * 80)
        lines.append("Migration Portion")
        lines.append("=" * 80)
        lines.append("")
        lines.extend(_format_route_rows(migration_portion))
        lines.append("")
    else:
        # Standard Add order
        lines.append("=" * 80)
        lines.append("SORTED ROUTE PATH (A -> B)")
        lines.append("=" * 80)
        lines.append("")
        lines.extend(_format_route_rows(items))
        lines.append("")

    return "\n".join(lines)


def format_notations(notations: list[str]) -> str:
    """Format the NOTATIONS section for display."""
    if not notations:
        return ""

    lines = [
        "",
        "=" * 80,
        "NOTATIONS",
        "=" * 80,
        "",
    ]
    for n in notations:
        lines.append(f"  {n}")
    lines.append("")
    return "\n".join(lines)


def _format_patching_lines(
    ticket: Ticket,
    has_stages: bool,
) -> list[str]:
    """Format patching lines for a single ticket.

    Handles standard add-order formatting, Stage 1 (HOT-CUT PREPARATION),
    and Stage 2 (HOT-CUT + CLEANUP) migration formatting.
    """
    lines: list[str] = []

    if has_stages and ticket.stage == 1:
        lines.append("Stage 1: HOT-CUT PREPARATION")
        lines.append("")
        _append_ticket_lines_with_site_breaks(
            lines,
            ticket.lines,
            _format_stage1_ticket_line,
        )

    elif has_stages and ticket.stage == 2:
        lines.append("Stage 2: HOT-CUT + CLEANUP")
        lines.append("")

        hotcut_lines = [
            tl
            for tl in ticket.lines
            if tl.hotcut_label in ("UNCHANGED", "NEW_ONLY")
            or (not tl.hotcut_label and tl.classification in ("LIVE", "NEW"))
        ]
        cleanup_lines = [
            tl
            for tl in ticket.lines
            if tl.hotcut_label == "DECOM_ONLY"
            or (not tl.hotcut_label and tl.classification == "DECOMMISSION")
        ]

        if hotcut_lines:
            lines.append("HOT-CUT (prep has been completed)")
            for tl in hotcut_lines:
                label = tl.hotcut_label if tl.hotcut_label else tl.classification
                if label in ("UNCHANGED", "LIVE"):
                    lines.append(f"{tl.text} *** HOT-CUT SIDE ***")
                else:
                    lines.append(tl.text)

        if cleanup_lines:
            if hotcut_lines:
                lines.append("")
            lines.append("CLEANUP (remove old patching)")
            _append_ticket_lines_with_site_breaks(lines, cleanup_lines)

    else:
        _append_ticket_lines_with_site_breaks(lines, ticket.lines)

    return lines


def _build_hub_details(
    ticket_sites: list[str],
    hub_lookup: dict[str, tuple[str, str]],
) -> str | None:
    """Build KEY IMPORTANT DETAILS content based on hub data.

    Returns None if section should be omitted (all sites are hubs).
    """
    from collections import defaultdict

    hub_to_satellites: dict[str, list[str]] = defaultdict(list)
    ship_lines: list[str] = []

    all_hub = True
    for site in ticket_sites:
        info = hub_lookup.get(site)
        if not info:
            continue  # Unknown sites handled by notations
        consumable, hub = info
        upper = (consumable or "").upper()
        if upper == "HUB":
            continue
        all_hub = False
        if upper == "SATELLITE" and hub:
            hub_to_satellites[hub].append(site)
        elif upper == "REMOTE" and hub:
            ship_lines.append(f"Consumables must be shipped from hub {hub} to {site}")

    if all_hub and not ship_lines:
        return None

    lines: list[str] = []
    for hub, satellites in sorted(hub_to_satellites.items()):
        lines.append(f"Collect cables for {', '.join(satellites)} from hub {hub}")
    lines.extend(ship_lines)

    return "\n".join(lines) if lines else None


def format_tickets(
    tickets: list[Ticket],
    all_planned: bool = False,
    bearer: str = "",
    service_id: str = "",
    hub_records: list[dict] | None = None,
) -> str:
    """Format field tech tickets for display.

    For migration orders with staged tickets:
    - Stage 1 (HOT-CUT PREPARATION): inline labels per line
    - Stage 2 (HOT-CUT + CLEANUP): HOT-CUT + CLEANUP sub-headers

    Args:
        tickets: List of Ticket objects.
        all_planned: True if all rows are Planned (hypothetical tickets).
        bearer: Bearer route path name (unused, kept for API compat).
        service_id: Service identifier (e.g., 'ICB-820729').
        hub_records: Optional hub records (unused, kept for API compat).

    Returns:
        Formatted string.
    """
    if not tickets:
        return "\n  (No active tickets - service not yet built)\n"

    has_stages = any(t.stage > 0 for t in tickets)
    parts: list[str] = []

    parts.append("")
    parts.append("=" * 80)
    parts.append("FIELD TECH TICKETS")
    parts.append("=" * 80)

    current_stage = -1
    for i, ticket in enumerate(tickets, 1):
        # Insert stage header when stage changes
        if has_stages and ticket.stage != current_stage:
            current_stage = ticket.stage
            parts.append("")
            if current_stage == 1:
                parts.append("=== STAGE 1: HOT-CUT PREPARATION ===")
            elif current_stage == 2:
                parts.append("=== STAGE 2: HOT-CUT + CLEANUP ===")

        parts.append("")
        parts.append(f"--- Ticket {i}: ---")
        if service_id:
            parts.append(f"Service: {service_id}")
        parts.append(f"Site: {ticket.cluster_name}")

        patching_lines = _format_patching_lines(ticket, has_stages)
        parts.append("")
        parts.extend(patching_lines)

    parts.append("")
    return "\n".join(parts)


_CONSOLIDATED_POC_AND_CHECKLIST = """\
=============

POC

PM operations contact
Use the approved project email channel
Use the approved project phone/WhatsApp channel

=============

LABELING AND CHECKLIST

Labeling Template:
- 1st row: IC#/ICB#
- 2nd row: Near-Side ODF/Equipment Location
- 3rd row: Far-Side ODF/Equipment Location

---

Checklist (Upload photos):
- Scope Images: Photos of passing fiber scope results for each connector installed
- Port Detail: Close-up of all ports/labels while connected (must be legible)
- Path and Management: Wide-angle shots showing jumpers properly dressed, routed, and separated
- Photos must prove fiber bend radius and strain relief are maintained end-to-end

Final Actions:
- Restore work area to data-center operational standards. If uncertain, ask data-center staff
- Notify POC before leaving site"""


def _build_consolidated_hub_lookup(
    hub_records: list[dict] | None,
) -> dict[str, tuple[str, str]]:
    """Return normalized hub metadata keyed by site code."""
    hub_lookup: dict[str, tuple[str, str]] = {}
    for record in hub_records or []:
        site_code = str(record.get("SITE_CODE", "")).strip()
        consumable = str(record.get("CONSUMABLE", "")).strip()
        hub = str(record.get("HUB", "")).strip()
        if site_code and (consumable or site_code not in hub_lookup):
            hub_lookup[site_code] = (consumable, hub)
    return hub_lookup


def _location_sites_for_hub_details(
    service_ticket_pairs: list[tuple[str, Ticket]],
) -> list[str]:
    """Collect unique non-stage-2 sites for consolidated hub guidance."""
    ordered_sites: list[str] = []
    seen_sites: set[str] = set()
    for _, ticket in service_ticket_pairs:
        if ticket.stage == 2:
            continue
        for site in ticket.sites:
            if site not in seen_sites:
                ordered_sites.append(site)
                seen_sites.add(site)
    return ordered_sites


def _append_consolidated_service_blocks(
    parts: list[str],
    service_ticket_pairs: list[tuple[str, Ticket]],
    has_stages: bool,
) -> None:
    """Append per-service patching blocks for one location."""
    for service_id, ticket in service_ticket_pairs:
        parts.append("")
        parts.append(f"- Service ID: {service_id}")
        parts.append(f"- Site: {ticket.cluster_name}")
        parts.append("")
        parts.append("Patching:")
        parts.extend(_format_patching_lines(ticket, has_stages))


def _append_consolidated_hub_details(
    parts: list[str],
    location_sites: list[str],
    hub_lookup: dict[str, tuple[str, str]],
) -> None:
    """Append the optional hub-details section for one location."""
    hub_details = _build_hub_details(location_sites, hub_lookup)
    if not hub_details:
        return

    parts.append("")
    parts.append("=============")
    parts.append("")
    parts.append("KEY IMPORTANT DETAILS")
    parts.append("")
    parts.append(hub_details)


def _append_consolidated_location_section(
    parts: list[str],
    location_id: str,
    service_ticket_pairs: list[tuple[str, Ticket]],
    hub_lookup: dict[str, tuple[str, str]],
) -> None:
    """Append one full consolidated location block."""
    parts.append("")
    parts.append(f"--- Location: {location_id} ---")

    has_stages = any(ticket.stage > 0 for _, ticket in service_ticket_pairs)
    _append_consolidated_service_blocks(parts, service_ticket_pairs, has_stages)
    _append_consolidated_hub_details(
        parts,
        _location_sites_for_hub_details(service_ticket_pairs),
        hub_lookup,
    )

    parts.append("")
    parts.append(_CONSOLIDATED_POC_AND_CHECKLIST)


def _write_cells(ws: Worksheet, row_num: int, values: Sequence[Any]) -> None:
    """Write one worksheet row of values starting at column 1."""
    for col_idx, value in enumerate(values, 1):
        ws.cell(row=row_num, column=col_idx, value=value)


def _route_sheet_row_values(item: InCARow) -> list[object]:
    """Return the legacy workbook row values for one route item."""
    points_out = item.display_points if item.display_points else item.cabling_points
    return [
        item.site_code,
        item.site_type,
        item.ne_info or "",
        item.cabling_location,
        points_out,
        item.conn_type,
        item.location_alias or "",
        item.route_path,
        item.pos,
        item.status_o_time or "",
        item.o_time or "",
        item.status_t_time or "",
        item.t_time or "",
        item.comment or "",
        item.classification,
    ]


def _write_route_data_rows(
    ws: Worksheet,
    start_row: int,
    row_items: list[InCARow],
) -> int:
    """Write route rows starting at start_row and return the next free row."""
    row_num = start_row
    for item in row_items:
        _write_cells(ws, row_num, _route_sheet_row_values(item))
        row_num += 1
    return row_num


def _write_route_headers(
    ws: Worksheet,
    row_num: int,
    headers: list[str],
) -> int:
    """Write the workbook route headers and return the next free row."""
    _write_cells(ws, row_num, headers)
    return row_num + 1


def _write_route_sheet_section(
    ws: Worksheet,
    row_num: int,
    title: str,
    headers: list[str],
    row_items: list[InCARow],
) -> int:
    """Write one labeled route section and return the next free row."""
    ws.cell(row=row_num, column=1, value=title)
    row_num = _write_route_headers(ws, row_num + 1, headers)
    return _write_route_data_rows(ws, row_num, row_items)


def _route_sheet_title(service_id: str, bearer: str) -> str:
    """Return the migration workbook title row."""
    title = f"Route path - {service_id}" if service_id else "Route path"
    return f"{title} | {bearer}" if bearer else title


def _write_sorted_route_sheet(
    ws_sorted: Worksheet,
    items: list[InCARow],
    migration_portion: list[InCARow] | None,
    service_id: str,
    bearer: str,
) -> None:
    """Write the Sorted Route Path sheet for standard and migration outputs."""
    ws_sorted.title = "Sorted Route Path"
    headers = INCA_COLUMNS + ["Classification"]

    if migration_portion is None:
        row_num = _write_route_headers(ws_sorted, 1, headers)
        _write_route_data_rows(ws_sorted, row_num, items)
        return

    ws_sorted.cell(row=1, column=1, value=_route_sheet_title(service_id, bearer))
    row_num = _write_route_sheet_section(ws_sorted, 2, "Current Route Path", headers, items)
    row_num += 1
    _write_route_sheet_section(ws_sorted, row_num, "Migration Portion", headers, migration_portion)


def _write_notations_sheet(
    ws_notes: Worksheet,
    notations: list[str],
) -> None:
    """Write the Notations sheet."""
    ws_notes.title = "Notations"
    ws_notes.cell(row=1, column=1, value="Notation")
    for row_num, note in enumerate(notations, 2):
        ws_notes.cell(row=row_num, column=1, value=note)


def _write_tickets_sheet(
    ws_tickets: Worksheet,
    tickets: list[Ticket],
) -> None:
    """Write the Tickets sheet."""
    ws_tickets.title = "Tickets"
    _write_cells(ws_tickets, 1, ["Ticket", "Variant", "Line"])

    ticket_row = 2
    for ticket in tickets:
        for ticket_line in ticket.lines:
            _write_cells(
                ws_tickets, ticket_row, [ticket.cluster_name, ticket_line.variant, ticket_line.text]
            )
            ticket_row += 1


def format_consolidated_tickets(
    location_tickets: dict[str, list[tuple[str, Ticket]]],
    hub_records: list[dict] | None = None,
) -> str:
    """Format consolidated field tech tickets grouped by physical location.

    Each location (identified by SITE_LOCATION_ID) gets a single consolidated
    ticket containing patching instructions from all services at that location,
    with shared POC/labeling/checklist and hub details appearing only once.

    Args:
        location_tickets: Mapping of SITE_LOCATION_ID (or fallback site_code)
            to list of (service_id, Ticket) pairs.
        hub_records: Optional HUB_SITE records for consumable hub details.

    Returns:
        Formatted string with consolidated ticket output.
    """
    if not location_tickets:
        return ""

    hub_lookup = _build_consolidated_hub_lookup(hub_records)
    parts: list[str] = []
    parts.append("")
    parts.append("=" * 80)
    parts.append("CONSOLIDATED FIELD TECH TICKETS")
    parts.append("=" * 80)

    for location_id in sorted(location_tickets.keys()):
        service_ticket_pairs = location_tickets[location_id]
        if not service_ticket_pairs:
            continue
        _append_consolidated_location_section(
            parts,
            location_id,
            service_ticket_pairs,
            hub_lookup,
        )

    parts.append("")
    return "\n".join(parts)


def write_output_excel(
    filepath: str,
    items: list[InCARow],
    notations: list[str],
    tickets: list[Ticket],
    migration_portion: list[InCARow] | None = None,
    service_id: str = "",
    bearer: str = "",
) -> None:
    """Write sorted route path and tickets to an Excel file.

    For migration orders (migration_portion is not None), emits section
    headers "Current Route Path" and "Migration Portion" with column
    headers repeated for each section.

    Args:
        filepath: Output .xlsx file path.
        items: Sorted InCARow objects (Current Route Path for migration).
        notations: NOTATIONS section entries.
        tickets: Generated tickets.
        migration_portion: Migration Portion rows (None for non-migration).
        service_id: Service identifier for title row.
        bearer: Bearer route path name for title row.
    """
    wb = openpyxl.Workbook()
    try:
        _write_sorted_route_sheet(
            cast(Worksheet, wb.active),
            items,
            migration_portion,
            service_id,
            bearer,
        )
        _write_notations_sheet(wb.create_sheet(), notations)
        _write_tickets_sheet(wb.create_sheet(), tickets)
        wb.save(filepath)
    finally:
        wb.close()
