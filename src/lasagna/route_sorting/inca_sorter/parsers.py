"""CSV reading, Snowflake parsing, edge parsing, and row construction."""

from __future__ import annotations

import csv
import json
import re
import sys
from collections import defaultdict
from collections.abc import Callable
from typing import Any, cast

import openpyxl

from .models import (
    INCA_COLUMNS,
    InCARow,
    SnowflakeCombinedData,
)

SNOWFLAKE_INCA_QIDS = {
    "TRUNK_ODF",
    "DEVICE",
    "ODUC",
    "DP_SDP",
    "EDGES",
    "TL_DEVICE",
    "HUB_SITE",
    "SITE_METADATA",
    "TRUNK_METADATA",
    "ROUTE_ORDER_METADATA",
    "TRANSMISSION_METADATA",
    "BO_FIBERS",
}

SDM_WORKSPACE_QIDS = {
    "00_COHORT_COUNT",
    "00_RUN_METADATA",
    "01_LEGACY_QUEUE",
    "02_ENRICHED_QUEUE",
    "04_SUMMARY",
    "05_ALL_WOS",
    "06_ALL_ASSIGNMENTS",
    "07_ALL_TPS",
    "08_FS_WO_DETAIL",
    "09B_SERVICE_FAMILY_DISCOVERY",
    "09_FS_COLUMN_DISCOVERY",
    "10_TPS_STATUS_HISTORY",
    "10_ASSIGNMENT_STATUS_HISTORY",
    "10_WO_STATUS_HISTORY",
    "10B_ALL_HISTORY_TABLES",
    "10_HISTORY_COLUMNS",
    "10_HISTORY_TABLE_DISCOVERY",
}

SDM_WORKSPACE_QID_PREFIXES = ("COMBINED_", "GEO-", "ASCOPE_", "ITEM")


def _is_sdm_workspace_qid(qid: str) -> bool:
    return qid in SDM_WORKSPACE_QIDS or qid.startswith(SDM_WORKSPACE_QID_PREFIXES)


def _safe_str(val: object) -> str:
    """Convert cell value to stripped string, handle None."""
    if val is None:
        return ""
    return str(val).strip()


def _safe_int(val: object) -> int:
    """Convert cell value to int, default 0."""
    if val is None:
        return 0
    if not isinstance(val, str | int | float):
        return 0
    try:
        return int(val)
    except (ValueError, TypeError):
        return 0


def _has_usable_cabling_points(val: str) -> bool:
    """Return True when an export row carries real BO-side patch points."""
    normalized = val.strip().upper()
    return bool(normalized and normalized not in {"NA", "N/A"})


def _csv_optional(val: object) -> str | None:
    """Convert value to str | None (None if empty/None)."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _split_device_location(device_location: str) -> tuple[str, str]:
    """Split DEVICE_LOCATION into (location_prefix, port_address).

    Handles two CCP.LOCATION formats from Snowflake:
    - Double-colon: '[building]rack::port:addr' -> ('[building]rack', 'port:addr')
    - Single-colon: '[building]rack:port' -> ('[building]rack', 'port')
    - No colon after bracket: returns (device_location, '')
    """
    bracket_end = device_location.rfind("]")
    if bracket_end < 0:
        return (device_location, "")
    after_bracket = device_location[bracket_end + 1 :]
    if "::" in after_bracket:
        parts = after_bracket.split("::", 1)
        return (device_location[: bracket_end + 1] + parts[0], parts[1])
    last_colon = after_bracket.rfind(":")
    if last_colon >= 0:
        return (
            device_location[: bracket_end + 1] + after_bracket[:last_colon],
            after_bracket[last_colon + 1 :],
        )
    return (device_location, "")


def _build_ne_information(
    ne: str,
    ne_part: str,
    optic_function: str,
    device_location: str,
    connection_point_nr: str,
    direction: str,
    chassis_function: str | None = None,
) -> str:
    """Construct INCA NE Information composite string from Snowflake fields.

    Formula: {NE} {NE_PART} -{CHASSIS}\\{OPTIC} -({PORT_ADDR}.{CPNR}:{DIR})

    Port address is extracted from device_location after '::' separator,
    with the last ':' replaced by '\\' to match INCA format.

    Args:
        ne: Device name (e.g., 'dls-b23' or 'DAL/C2 XS G40 24').
        ne_part: Device part/chassis (e.g., 'NCS-5508' or 'G42 01').
        optic_function: Optic function (e.g., 'QDD-400G-LR4-S').
        device_location: Full CCP.LOCATION (e.g., '[BLDG]rack::port:addr').
        connection_point_nr: Connection point number (e.g., '01' or '.').
        direction: 'Tx' or 'Rx'.
        chassis_function: Optional chassis function from Query C (e.g., 'CHM6-C8').
            Falls back to ne_part if not provided.

    Returns:
        Composite NE Information string matching INCA export format.
    """
    # Extract port address from device location
    _, port_addr = _split_device_location(device_location)
    # Replace last ':' with '\' to match INCA format (only for :: format)
    if "::" in device_location:
        last_colon = port_addr.rfind(":")
        if last_colon >= 0:
            port_addr = port_addr[:last_colon] + "\\" + port_addr[last_colon + 1 :]

    chassis = chassis_function if chassis_function else ne_part
    return (
        f"{ne} {ne_part} -{chassis}\\{optic_function} "
        f"-({port_addr}.{connection_point_nr}:{direction})"
    )


def _parse_cabling_point(raw: str) -> str | None:
    """Extract the numeric cabling point from INCA format.

    '45 Cable.45' -> '45'
    '05 Cable.05' -> '05' (preserve leading zeros as-is from INCA)
    'N/A' or 'NA' -> None
    """
    raw = raw.strip()
    if not raw or raw.upper() in ("N/A", "NA"):
        return None
    # Take the first token (number before ' Cable.')
    m = re.match(r"(\d+)", raw)
    if m:
        return m.group(1)
    return raw


def _cabling_point_int(raw: str) -> int:
    """Extract numeric cabling point as integer for sort ordering."""
    raw = raw.strip()
    if not raw or raw.upper() in ("N/A", "NA"):
        return 0
    m = re.match(r"(\d+)", raw)
    return int(m.group(1)) if m else 0


def _ne_group_key(ne_info: str | None) -> str:
    """Extract device grouping key from NE Information string.

    Returns '{NE} {NE_PART}' prefix (before first ' -'), which is
    identical between INCA and Snowflake sources regardless of
    chassis/linecard model differences.
    """
    if not ne_info or not ne_info.strip():
        return "unknown"
    return ne_info.strip().split(" -")[0]


_PLANNED_DEVICE_LOCATION = "NE-location: (planned device, not yet installed)"
_MISSING_DEVICE_LOCATION = "NE-location: (no BO ODF location in INCA)"
_MISSING_BO_ODF_LOCATION = "BO ODF location missing from Snowflake export"


def _make_snowflake_trunk_row(rec: dict, row_idx: int, service_id: str | None = None) -> InCARow:
    """Create an InCARow from a Snowflake trunk ODF record."""
    row = InCARow(
        site_code=_safe_str(rec.get("SITE_CODE")),
        site_type=_safe_str(rec.get("SITE_TYPE")),
        ne_info=None,
        cabling_location=_safe_str(rec.get("CABLING_LOCATION")),
        cabling_points=_safe_str(rec.get("CABLING_POINTS")),
        conn_type=_safe_str(rec.get("CONN_TYPE")),
        location_alias=_csv_optional(rec.get("LOCATION_ALIAS")),
        route_path=_safe_str(rec.get("ROUTE_PATH")),
        pos=_safe_int(rec.get("POS", 0)),
        status_o_time=_csv_optional(rec.get("STATUS_O_TIME")),
        o_time=_csv_optional(rec.get("O_TIME")),
        status_t_time=_csv_optional(rec.get("STATUS_T_TIME")),
        t_time=_csv_optional(rec.get("T_TIME")),
        comment=_csv_optional(rec.get("COMMENT")),
        row_index=row_idx,
        service_id=service_id,
        dp_owner=None,
        site_type_no=_safe_str(rec.get("SITE_TYPE_NO")),
        # Phase 1: structured location fields from CCP
        floor=_safe_str(rec.get("FLOOR")),
        room=_safe_str(rec.get("ROOM")),
        row=_safe_str(rec.get("ROW_")),
        rowside=_safe_str(rec.get("ROWSIDE")),
        rack=_safe_str(rec.get("RACK")),
        shelf=_safe_str(rec.get("SHELF")),
        subrack=_safe_str(rec.get("SUBRACK")),
        connection_point_nr=_safe_str(rec.get("CONNECTION_POINT_NR")),
    )
    row.site_side = _csv_optional(rec.get("SITE_SIDE"))
    return row


def _location_has_meaningful_building_data(location_prefix: str) -> bool:
    """Return True when a location prefix includes more than punctuation or digits."""
    loc_check = location_prefix
    for ch in "/.:\\- 0123456789":
        loc_check = loc_check.replace(ch, "")
    return bool(loc_check.strip())


def _resolve_device_ne_location_label(location: str) -> str | None:
    """Return an NE-location label when Snowflake provides usable building data."""
    location_prefix, _ = _split_device_location(location)
    if not _location_has_meaningful_building_data(location_prefix):
        return None
    return f"NE-location: {location_prefix}"


def _fallback_device_ne_location(rec: dict, device_location: str) -> str:
    """Resolve the owner-readable NE-location fallback for a Snowflake device row."""
    for location in (device_location, _safe_str(rec.get("NEP_LOCATION"))):
        label = _resolve_device_ne_location_label(location)
        if label:
            return label

    status = _safe_str(rec.get("STATUS_O_TIME"))
    if status.lower() == "planned":
        return _PLANNED_DEVICE_LOCATION
    return _MISSING_DEVICE_LOCATION


def _resolve_device_cabling_fields(rec: dict, device_location: str) -> tuple[str, str, str]:
    """Resolve cabling location, points, and connection type for a device row."""
    cabling_location = _safe_str(rec.get("CABLING_LOCATION"))
    if cabling_location:
        return (
            cabling_location,
            _safe_str(rec.get("CABLING_POINTS")),
            _safe_str(rec.get("CONN_TYPE")),
        )

    cabling_points = _safe_str(rec.get("CABLING_POINTS"))
    conn_type = _safe_str(rec.get("CONN_TYPE"))
    if _has_usable_cabling_points(cabling_points) and conn_type:
        return (_MISSING_BO_ODF_LOCATION, cabling_points, conn_type)

    return (_fallback_device_ne_location(rec, device_location), "NA", "")


def _make_snowflake_device_row(
    rec: dict,
    row_idx: int,
    chassis_lookup: dict[tuple[str, str], str],
    service_id: str | None = None,
) -> InCARow:
    """Create an InCARow from a Snowflake device record.

    Builds NE Information composite string, detects NE-Location rows,
    and looks up chassis function from ODUC records.
    """
    site_code = _safe_str(rec.get("SITE_CODE"))
    ne = _safe_str(rec.get("NE"))
    ne_part = _safe_str(rec.get("NE_PART"))
    optic_function = _safe_str(rec.get("OPTIC_FUNCTION"))
    device_location = _safe_str(rec.get("DEVICE_LOCATION"))
    connection_point_nr = _safe_str(rec.get("CONNECTION_POINT_NR"))
    direction = _safe_str(rec.get("DIRECTION"))

    chassis = chassis_lookup.get((site_code, ne))
    ne_info = _build_ne_information(
        ne,
        ne_part,
        optic_function,
        device_location,
        connection_point_nr,
        direction,
        chassis,
    )
    cabling_location, cabling_points, conn_type = _resolve_device_cabling_fields(
        rec, device_location
    )

    return InCARow(
        site_code=site_code,
        site_type=_safe_str(rec.get("SITE_TYPE")),
        ne_info=ne_info,
        cabling_location=cabling_location,
        cabling_points=cabling_points,
        conn_type=conn_type,
        location_alias=_csv_optional(rec.get("LOCATION_ALIAS")),
        route_path=_safe_str(rec.get("ROUTE_PATH")),
        pos=_safe_int(rec.get("POS", 0)),
        status_o_time=_csv_optional(rec.get("STATUS_O_TIME")),
        o_time=_csv_optional(rec.get("O_TIME")),
        status_t_time=_csv_optional(rec.get("STATUS_T_TIME")),
        t_time=_csv_optional(rec.get("T_TIME")),
        comment=_csv_optional(rec.get("COMMENT")),
        row_index=row_idx,
        service_id=service_id,
        dp_owner=None,
        site_type_no=_safe_str(rec.get("SITE_TYPE_NO")),
        # Phase 1: structured fields from CCP/NE_PART
        # Device rows use BO_ prefixed columns for the resolved BO ODF location
        floor=_safe_str(rec.get("BO_FLOOR", rec.get("FLOOR", ""))),
        room=_safe_str(rec.get("BO_ROOM", rec.get("ROOM", ""))),
        row=_safe_str(rec.get("BO_ROW", rec.get("ROW_", ""))),
        rowside=_safe_str(rec.get("BO_ROWSIDE", rec.get("ROWSIDE", ""))),
        rack=_safe_str(rec.get("BO_RACK", rec.get("RACK", ""))),
        shelf=_safe_str(rec.get("BO_SHELF", rec.get("SHELF", ""))),
        subrack=_safe_str(rec.get("BO_SUBRACK", rec.get("SUBRACK", ""))),
        connection_point_nr=connection_point_nr,
        ne_type=_safe_str(rec.get("NE_TYPE")),
        ne_function=_safe_str(rec.get("NE_FUNCTION")),
        # Phase 2A: structured port assembly fields
        slot=_safe_str(rec.get("SLOT")),
        subslot=_safe_str(rec.get("SUBSLOT")),
    )


_DEMARC_POS_SUFFIX = re.compile(r"\s+pos\s+\d+$", re.IGNORECASE)


def _dp_sdp_ne_information(rec: dict) -> str | None:
    """Return demarcation NE information using the physical DP/SDP function."""
    raw = _safe_str(rec.get("NE_INFORMATION"))
    function = _safe_str(rec.get("FUNCTION"))
    if not raw or not function:
        return raw or None
    parts = raw.split()
    if len(parts) >= 2 and parts[0] in {"DP", "SDP"}:
        return f"{parts[0]} {function}"
    return raw


def _make_snowflake_dp_sdp_row(rec: dict, row_idx: int, service_id: str | None = None) -> InCARow:
    """Create an InCARow from a Snowflake DP/SDP demarcation record.

    DP_SDP rows have NE_INFORMATION pre-computed as 'DP ODF' or 'SDP ODF',
    and ROUTE_PATH prefixed with 'Demarcation point: ...'.
    """
    # Legacy CSVs leak a trailing ' pos <int>' onto demarc ROUTE_PATH; strip
    # only that exact suffix so old exports render the same demarc text as
    # post-SQL-fix exports. Cannot reconstruct customer/name fields when the
    # source CSV does not carry NWP_ID / NWP_CUSTOMER as separate columns.
    route_path = _DEMARC_POS_SUFFIX.sub("", _safe_str(rec.get("ROUTE_PATH")))
    return InCARow(
        site_code=_safe_str(rec.get("SITE_CODE")),
        site_type=_safe_str(rec.get("SITE_TYPE")),
        ne_info=_dp_sdp_ne_information(rec),
        cabling_location=_safe_str(rec.get("CABLING_LOCATION")),
        cabling_points=_safe_str(rec.get("CABLING_POINTS")),
        conn_type=_safe_str(rec.get("CONN_TYPE")),
        location_alias=_csv_optional(rec.get("LOCATION_ALIAS")),
        route_path=route_path,
        pos=_safe_int(rec.get("POS", 0)),
        status_o_time=_csv_optional(rec.get("STATUS_O_TIME")),
        o_time=_csv_optional(rec.get("O_TIME")),
        status_t_time=_csv_optional(rec.get("STATUS_T_TIME")),
        t_time=_csv_optional(rec.get("T_TIME")),
        comment=_csv_optional(rec.get("COMMENT")),
        row_index=row_idx,
        service_id=service_id,
        dp_owner=_csv_optional(rec.get("DP_OWNER")),
        site_type_no=_safe_str(rec.get("SITE_TYPE_NO")),
    )


def _normalize_bearer_endpoint_device_site_types(
    services: dict[str, list[InCARow]],
) -> None:
    """Normalize logical bearer BR endpoint devices to physical XS rows.

    Some Snowflake DEVICE rows carry the bearer marker (`BR`) instead of the
    accessible physical site type. When the same service and site also has XS
    ODF evidence, the endpoint device belongs with that XS field-tech block.
    Without this, sorted route paths show a stray BR block and ticket generation
    drops the direct endpoint patch because tickets are XS-only.
    """
    for rows in services.values():
        xs_sites = {r.site_code for r in rows if r.site_type == "XS"}
        if not xs_sites:
            continue
        for row in rows:
            if (
                row.site_type == "BR"
                and row.site_code in xs_sites
                and row.is_device_row
                and row.is_router
            ):
                row.site_type = "XS"
                row.site_type_no = ""


def _warn_unexpected_combined_header(header: list[str]) -> None:
    """Warn when a combined CSV uses a non-canonical header row."""
    header_upper = [h.strip().upper() for h in header]
    if header_upper == ["QID", "ROW_DATA"]:
        return
    print(
        f"WARNING: Unexpected CSV header {header}, expected ['QID', 'ROW_DATA']",
        file=sys.stderr,
    )


def _parse_combined_row_data(raw_json: str, qid: str, line_num: int) -> dict[str, Any] | None:
    """Parse a combined CSV ROW_DATA cell, preserving owner-readable warnings."""
    if not raw_json:
        print(
            f"WARNING: Empty ROW_DATA at line {line_num} (QID={qid}), skipping",
            file=sys.stderr,
        )
        return None
    try:
        record = json.loads(raw_json)
    except json.JSONDecodeError as exc:
        print(
            f"WARNING: Invalid JSON at line {line_num} (QID={qid}): {exc}",
            file=sys.stderr,
        )
        return None
    if not isinstance(record, dict):
        print(
            f"WARNING: ROW_DATA at line {line_num} (QID={qid}) is not an object, skipping",
            file=sys.stderr,
        )
        return None
    return record


def _group_combined_record(
    by_qid: dict[str, list[dict]],
    skipped_qids: set[str],
    qid: str,
    rec: dict,
) -> None:
    """Store a parsed combined CSV record in the correct QID bucket."""
    if qid in SNOWFLAKE_INCA_QIDS:
        by_qid[qid].append(rec)
        return
    if not _is_sdm_workspace_qid(qid):
        skipped_qids.add(qid)


def _read_combined_qid_groups(filepath: str) -> tuple[dict[str, list[dict]], set[str]] | None:
    """Read and group combined CSV records by QID."""
    by_qid: dict[str, list[dict]] = defaultdict(list)
    skipped_qids: set[str] = set()

    with open(filepath, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        header = next(reader, None)
        if header is None:
            print("WARNING: Empty combined CSV file", file=sys.stderr)
            return None

        _warn_unexpected_combined_header(header)

        for line_num, csv_row in enumerate(reader, start=2):
            if len(csv_row) < 2:
                continue
            qid = csv_row[0].strip()
            rec = _parse_combined_row_data(csv_row[1].strip(), qid, line_num)
            if rec is None:
                continue
            _group_combined_record(by_qid, skipped_qids, qid, rec)

    return by_qid, skipped_qids


def _build_combined_chassis_lookup(oduc_records: list[dict]) -> dict[tuple[str, str], str]:
    """Build the combined-export chassis lookup from ODUC records."""
    chassis_lookup: dict[tuple[str, str], str] = {}
    for rec in oduc_records:
        site_code = _safe_str(rec.get("SITE_CODE"))
        ne = _safe_str(rec.get("NE"))
        chassis_fn = _safe_str(rec.get("CHASSIS_FUNCTION"))
        if site_code and ne and chassis_fn:
            chassis_lookup[(site_code, ne)] = chassis_fn
    return chassis_lookup


def _append_combined_service_rows(
    services: dict[str, list[InCARow]],
    records: list[dict],
    row_idx: int,
    row_factory: Callable[[dict, int, str], InCARow],
) -> int:
    """Append service-scoped combined CSV records in row-construction order."""
    for rec in records:
        service_id = _safe_str(rec.get("SERVICE_ID"))
        if not service_id:
            continue
        services[service_id].append(row_factory(rec, row_idx, service_id))
        row_idx += 1
    return row_idx


def _build_combined_services(
    by_qid: dict[str, list[dict]],
    chassis_lookup: dict[tuple[str, str], str],
) -> dict[str, list[InCARow]]:
    """Build combined-export InCARow groups keyed by service ID."""
    services: dict[str, list[InCARow]] = defaultdict(list)
    row_idx = 1
    row_idx = _append_combined_service_rows(
        services,
        by_qid.get("TRUNK_ODF", []),
        row_idx,
        _make_snowflake_trunk_row,
    )
    row_idx = _append_combined_service_rows(
        services,
        by_qid.get("DEVICE", []),
        row_idx,
        lambda rec, idx, service_id: _make_snowflake_device_row(
            rec,
            idx,
            chassis_lookup,
            service_id=service_id,
        ),
    )
    _append_combined_service_rows(
        services,
        by_qid.get("DP_SDP", []),
        row_idx,
        _make_snowflake_dp_sdp_row,
    )
    return dict(services)


def read_snowflake_csv(
    query_a_path: str,
    query_b_path: str | None = None,
    query_c_path: str | None = None,
) -> list[InCARow]:
    """Read Snowflake CSV exports and construct InCARow objects.

    Reads Query A (trunk ODF rows) and Query B (device rows with cable trace),
    optionally Query C (ODUC chassis function), and returns a combined list of
    InCARow objects compatible with the downstream sorting pipeline.

    Args:
        query_a_path: Path to Query A CSV (trunk ODF rows).
        query_b_path: Path to Query B CSV (device rows). Optional.
        query_c_path: Path to Query C CSV (ODUC chassis function). Optional.

    Returns:
        Combined list of InCARow objects from both queries.
    """
    rows: list[InCARow] = []
    row_idx = 1

    chassis_lookup: dict[tuple[str, str], str] = {}
    if query_c_path:
        with open(query_c_path, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                key = (row["SITE_CODE"].strip(), row["NE"].strip())
                chassis_lookup[key] = row["CHASSIS_FUNCTION"].strip()

    with open(query_a_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if not row.get("SITE_CODE", "").strip():
                continue
            r = _make_snowflake_trunk_row(row, row_idx)
            rows.append(r)
            row_idx += 1

    if query_b_path:
        with open(query_b_path, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                if not row.get("SITE_CODE", "").strip():
                    continue
                r = _make_snowflake_device_row(row, row_idx, chassis_lookup)
                rows.append(r)
                row_idx += 1

    return rows


def read_snowflake_combined_csv(filepath: str) -> SnowflakeCombinedData:
    """Read combined Snowflake CSV export (QID,ROW_DATA format).

    The SQL produces: SELECT qid, row_data FROM prod_all
    QID values: TRUNK_ODF, DEVICE, ODUC, DP_SDP, EDGES
    ROW_DATA: JSON object with query-specific fields.

    Processing order: ODUC first (builds chassis_lookup), then TRUNK_ODF,
    DEVICE (uses chassis_lookup), DP_SDP, and EDGES (raw dicts for hierarchy).

    Returns:
        SnowflakeCombinedData with services dict and edge_records list.
    """
    grouped_rows = _read_combined_qid_groups(filepath)
    if grouped_rows is None:
        return SnowflakeCombinedData(services={}, edge_records=[], tl_device_records=[])
    by_qid, skipped_qids = grouped_rows

    if skipped_qids:
        print(
            f"WARNING: Skipped unknown QID values: {sorted(skipped_qids)}",
            file=sys.stderr,
        )

    chassis_lookup = _build_combined_chassis_lookup(by_qid.get("ODUC", []))
    services = _build_combined_services(by_qid, chassis_lookup)

    # EDGES rows (raw dicts for hierarchy edge parsing)
    edge_records = by_qid.get("EDGES", [])

    # TL_DEVICE rows (raw dicts for transport-link-to-device mapping)
    tl_device_records = by_qid.get("TL_DEVICE", [])

    # HUB_SITE / SITE_METADATA rows (site metadata including hub mapping)
    # SITE_METADATA is the new QID; HUB_SITE is the legacy QID for backward compat
    hub_records = by_qid.get("SITE_METADATA", []) or by_qid.get("HUB_SITE", [])

    # TRUNK_METADATA rows (PCG metadata for trunk endpoint/type resolution)
    trunk_metadata = by_qid.get("TRUNK_METADATA", [])

    # ROUTE_ORDER_METADATA rows (Snowflake route position contract)
    route_order_metadata = by_qid.get("ROUTE_ORDER_METADATA", [])

    # TRANSMISSION_METADATA rows (transmission metadata for TL edge endpoint resolution)
    transmission_metadata = by_qid.get("TRANSMISSION_METADATA", [])

    # BO_FIBERS rows (breakout fiber traces for notation enrichment)
    bo_fibers = by_qid.get("BO_FIBERS", [])

    _normalize_bearer_endpoint_device_site_types(services)

    return SnowflakeCombinedData(
        services=services,
        edge_records=edge_records,
        tl_device_records=tl_device_records,
        hub_records=hub_records,
        trunk_metadata=trunk_metadata,
        route_order_metadata=route_order_metadata,
        transmission_metadata=transmission_metadata,
        bo_fibers=bo_fibers,
    )


def extract_service_id(filepath: str) -> str | None:
    """Extract service ID from Excel file title row.

    INCA exports have a title in row 1, cell A1:
    'Route path with cabling points - IC-136025 | ...'
    'Route path with cabling points - ICB-811386 | ...'

    Returns the service ID (e.g., 'IC-136025', 'ICB-811386') or None.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = cast(Any, wb.active)
    row1 = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    wb.close()
    if not row1 or not row1[0]:
        return None
    title = str(row1[0][0]) if row1[0][0] else ""
    m = re.search(r"(ICB?-\d+)", title)
    return m.group(1) if m else None


def _map_excel_headers_case_insensitively(
    row_cells: tuple[Any, ...],
    inca_col_lower: dict[str, str],
) -> dict[str, int]:
    """Map case-insensitive Excel headers to canonical INCA column names."""
    header_map: dict[str, int] = {}
    for cell in row_cells:
        header = _safe_str(cell.value).lower()
        if header in inca_col_lower:
            header_map[inca_col_lower[header]] = cell.column - 1
    return header_map


def _row_contains_site_code_header(row_cells: tuple[Any, ...]) -> bool:
    """Return True when a worksheet row includes the Site Code header."""
    return any(_safe_str(cell.value).lower() == "site code" for cell in row_cells)


def _fallback_excel_header_map(ws: Any) -> tuple[int, dict[str, int]]:
    """Fallback to row 1 headers for synthetic workbook fixtures."""
    header_map: dict[str, int] = {}
    first_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=False))
    if first_row and first_row[0]:
        for cell in first_row[0]:
            value = _safe_str(cell.value)
            if value in INCA_COLUMNS:
                header_map[value] = cell.column - 1
    return 1, header_map


def _find_excel_header_row(ws: Any, inca_col_lower: dict[str, str]) -> tuple[int, dict[str, int]]:
    """Find the Excel header row and its canonical INCA column map."""
    for row_num in range(1, 11):
        row_cells = list(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=False))
        if not row_cells or not row_cells[0]:
            continue
        cells = row_cells[0]
        if _row_contains_site_code_header(cells):
            return row_num, _map_excel_headers_case_insensitively(cells, inca_col_lower)
    return _fallback_excel_header_map(ws)


def _warn_missing_excel_columns(header_map: dict[str, int]) -> None:
    """Emit the existing owner-readable warning for absent INCA columns."""
    missing = set(INCA_COLUMNS) - set(header_map)
    if missing:
        print(f"WARNING: Missing columns in Excel: {missing}", file=sys.stderr)


def _excel_cell_value(row_data: tuple, header_map: dict[str, int], col_name: str) -> str:
    """Return a normalized Excel cell value for a canonical INCA column."""
    column_index = header_map.get(col_name)
    if column_index is None or column_index >= len(row_data):
        return ""
    return _safe_str(row_data[column_index])


def _make_excel_row(row_data: tuple, row_index: int, header_map: dict[str, int]) -> InCARow:
    """Create an InCARow from one Excel worksheet data row."""
    pos_value = row_data[header_map["Pos"]] if "Pos" in header_map else None
    return InCARow(
        site_code=_excel_cell_value(row_data, header_map, "Site Code"),
        site_type=_excel_cell_value(row_data, header_map, "Site Type"),
        ne_info=_excel_cell_value(row_data, header_map, "NE Information") or None,
        cabling_location=_excel_cell_value(row_data, header_map, "Cabling Location"),
        cabling_points=_excel_cell_value(row_data, header_map, "Cabling Points"),
        conn_type=_excel_cell_value(row_data, header_map, "Conn type"),
        location_alias=_excel_cell_value(row_data, header_map, "Location Alias") or None,
        route_path=_excel_cell_value(row_data, header_map, "Route Path"),
        pos=_safe_int(pos_value),
        status_o_time=_excel_cell_value(row_data, header_map, "Status o-time") or None,
        o_time=_excel_cell_value(row_data, header_map, "O-time") or None,
        status_t_time=_excel_cell_value(row_data, header_map, "Status t-time") or None,
        t_time=_excel_cell_value(row_data, header_map, "T-time") or None,
        comment=_excel_cell_value(row_data, header_map, "Comment") or None,
        row_index=row_index,
    )


def _read_excel_rows(ws: Any, data_start: int, header_map: dict[str, int]) -> list[InCARow]:
    """Read all non-blank Excel data rows into InCARow objects."""
    rows: list[InCARow] = []
    for row_index, row_data in enumerate(
        ws.iter_rows(min_row=data_start, values_only=True), start=data_start
    ):
        if all(value is None for value in row_data):
            continue
        row = _make_excel_row(row_data, row_index, header_map)
        if row.site_code:
            rows.append(row)
    return rows


def read_excel(filepath: str) -> list[InCARow]:
    """Read INCA export from Excel file, return list of InCARow objects.

    Handles real INCA exports where headers may not be on row 1.
    Scans rows 1-10 for the header row (case-insensitive column matching).
    Extra columns in the export are ignored gracefully.

    Args:
        filepath: Path to .xlsx file with INCA route path export.

    Returns:
        List of InCARow objects, one per data row.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    try:
        ws = cast(Any, wb.active)
        inca_col_lower = {col.lower(): col for col in INCA_COLUMNS}
        header_row_num, header_map = _find_excel_header_row(ws, inca_col_lower)
        _warn_missing_excel_columns(header_map)
        return _read_excel_rows(ws, header_row_num + 1, header_map)
    finally:
        wb.close()


def build_tl_device_map(
    tl_device_records: list[dict],
    service_id: str,
) -> dict[tuple[str, str], dict[str, list[str]]]:
    """Build a mapping from (service_id, site_code) to {tl_name: [ne_parts]}.

    This maps transport link names to the specific device(s) (NE_PART) that
    terminate them at each site, enabling data-driven within-site ordering.
    A single TL can have multiple NE_PARTs at one site (e.g., both XT-05 and
    OTC-02 at the same site for the same transport link).

    Args:
        tl_device_records: Raw dicts from TL_DEVICE rows in combined CSV.
        service_id: Filter to this service only.

    Returns:
        Dict keyed by (service_id, site_code) -> {tl_name: [ne_part, ...]}.
        Example: {('IC-394531', 'ATM/2'): {'ASH/2-ATM/2 OCGX05': ['XTC-08'],
                                             'ATM/2-IPLS O600G03': ['XT-44']}}
    """
    result: dict[tuple[str, str], dict[str, list[str]]] = defaultdict(lambda: defaultdict(list))
    for rec in tl_device_records:
        sid = str(rec.get("SERVICE_ID", "")).strip()
        if sid != service_id:
            continue
        tl_name = str(rec.get("TL_NAME", "")).strip()
        site_code = str(rec.get("SITE_CODE", "")).strip()
        ne_part = str(rec.get("NE_PART", "")).strip()
        if tl_name and site_code and ne_part:
            if ne_part not in result[(sid, site_code)][tl_name]:
                result[(sid, site_code)][tl_name].append(ne_part)
    return {k: dict(v) for k, v in result.items()}
